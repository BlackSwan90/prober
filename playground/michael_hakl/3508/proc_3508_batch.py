import os
import glob
import astropy.io.fits as fits
import numpy as np
import matplotlib.pyplot as plt
from scipy.ndimage import measurements as ms
import re
import pickle
import openpyxl
from imageio import imwrite
from PIL import Image as im
from matplotlib.ticker import IndexLocator
import xmltodict as xd
import dpath.util as du
from contextlib import suppress


D = {}
D['xfile_list'] = []
D['root_list'] = []

# #w7
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-07.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-034759-07_new')

#w7-old
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-07_MHL.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043753-07')

#w10-old
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-10_MHL.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043753-10\3508-043753-10')
# #
# # #w11-old
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-11_MHL.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043753-11\3508-043753-11')
#

#w12_3
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-12_3\3508-043759-12_3')

#w12
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-12_new.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-12_new\3508-043759-12_n4')
# #
# # #w12-old
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-12_MHL.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-12')
# #
# #w13
# D['xfile_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-13.xlsx')
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-13\3508-043759-13_new1')

#w13_3
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\3508-043759-13_3\3508-043759-13_3')

#B10380-w1
# D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\NL10380-01')

#B10380-w2
#D['root_list'].append(r'c:\Users\michael.hakl\Documents\XRO3508\NL10380-02\NL10380-02')

#B10380-w3
D['root_list'].append(r'W:\Production\DATA\WAM-######\WAM-003508\BATCH_3508-10380\NL10380-03\NL10380-03')


def get_map_wafer(d):
    d['N_r'] = 24
    d['N_c'] = 4
    map = np.zeros((d['N_r'], d['N_c']))
    map[5:20, 0] = 1
    map[:, 1] = 1
    map[:, 2] = 1
    map[5:20,3] = 1

    d['map_wafer'] = map
    return d

def load_im(d):

    #predelete empty dirs from list
    for i, die in enumerate(d['dirs_die']):
        if len( glob.glob(os.path.join(d['dirs_die'][i], 'ROIC 0','*.fits')) ) == 0:
            del d['dirs_die'][i]
    d['dirs_die'] = np.array(d['dirs_die'])
    d['N_die'] = len(d['dirs_die'])

    d['im_inj'] = np.zeros((d['N_die'],d['N_gain'],1032))
    d['im_dark'] = np.zeros((d['N_die'],d['N_gain'],1032))
    d['im_shortint'] = np.zeros((d['N_die'],d['N_gain'],1032))
    d['im_longint'] = np.zeros((d['N_die'], d['N_gain'], 1032))

    die_row = []
    die_col = []
    for i, die in enumerate(d['dirs_die']):
        die_coor = re.findall("R..-C..", die)
        die_row.append(int(re.findall("(\d\d)", die_coor[0])[0]))
        die_col.append(int(re.findall("(\d\d)", die_coor[0])[1]))

    d['map_RC_notation'] = np.zeros((d['N_r'], d['N_c']), dtype='<U16')
    d['map_RC_permut'] = np.full((d['N_r'], d['N_c']), fill_value=-1, dtype='int16')
    for i in range(d["N_die"]):
        d['map_RC_notation'][die_row[i] - 1, die_col[i] - 1] = 'R' + '{:0>2}'.format(die_row[i]) + 'C' + '{:0>2}'.format(die_col[i])
        d['map_RC_permut'][die_row[i] - 1, die_col[i] - 1] = i  # log the order of filling the wafer with folder labels
    d['map_RC_notation'] = d['map_RC_notation'][::-1, :]

    d['map_RC_permut'] = d['map_RC_permut'][::-1, :]  # reorder upside down
    permut_flat = np.ravel(d['map_RC_permut']) #flatten into 1d
    d['permut_flat_noneg'] = permut_flat[permut_flat != -1] #from rectangular to circular wafer
    d['dirs_die_sorted'] = d['dirs_die'][d['permut_flat_noneg'].astype(int)]  # reshuffle dirs to normal ordering
    
    i_die = 0
    d['map_dirs_die'] = np.full((d["N_r"],d["N_c"]),fill_value=-1,dtype='int16')
    for i in range(d["N_r"]):
        for ii in range(d["N_c"]):
            if d['map_RC_notation'][i,ii] != '': #skip if no die exist
                d['map_dirs_die'][i,ii] = i_die
                i_die += 1

    d['path_to_die']= [[],]*d['N_die']
    for i_die in range(d["N_die"]):
        d['path_to_die'][i_die] = os.path.join(d['dirs_die_sorted'][i_die], 'ROIC 0')
        for i_gain, gain in enumerate(d['list_gain']):
            path_to_gain = os.path.join(d['path_to_die'][i_die],gain)
            list_f_gain = [f for f in glob.glob(path_to_gain)]

            path_dark = [f for f in list_f_gain if 'DarkImage' in f]
            d['im_dark'][i_die, i_gain, :] = fits.open(path_dark[0])[0].data.squeeze()

            path_sat = [f for f in list_f_gain if 'SaturatedImage' in f]
            d['im_inj'][i_die, i_gain, :] = fits.open(path_sat[0])[0].data.squeeze() # sequence: element in filelist, hdu primary in fits, array in hdu, make row + remove redundant dimension

            path_shortint = [f for f in list_f_gain if 'ShortIntImage' in f]
            d['im_shortint'][i_die, i_gain, :] = fits.open(path_shortint[0])[0].data.squeeze()

            path_longint = [f for f in list_f_gain if 'LongIntImage' in f]
            d['im_longint'][i_die, i_gain, :] = fits.open(path_longint[0])[0].data.squeeze()

    d['im_DC'] = np.abs(d['im_longint']) - np.abs(d['im_shortint'])

    return d

def load_curr(d):
    d['curr_poweron'] = np.zeros((d['N_die'], 19))
    d['curr_CBGB'] = np.zeros((d['N_die'], 19))
    d['curr_running'] = np.zeros((d['N_die'],19))
    for i_die, die in enumerate(d['dirs_die_sorted']):
        f_xml = glob.glob(os.path.join(die,'*Current Measurements.xml'))
        f_xml = open(f_xml[0],'r')
        xml = f_xml.read()
        d_xml = xd.parse(xml)
        curr_poweron_str = du.get(d_xml,'measurement/iteration1/measurementData/PowerConsumptionResultPoweron/PowerConsumptionResult/tr/1/@trData')
        curr_CBGB_str = du.get(d_xml,'measurement/iteration1/measurementData/PowerConsumptionResultCBGBStarted/PowerConsumptionResult/tr/1/@trData')
        curr_running_str = du.get(d_xml,'measurement/iteration1/measurementData/PowerConsumptionResultRunningMode/PowerConsumptionResult/tr/1/@trData')
        d['curr_running'][i_die,:] = np.array(curr_running_str.split(';')[1:],dtype='float64')
        d['curr_CBGB'][i_die, :] = np.array(curr_CBGB_str.split(';')[1:], dtype='float64')
        d['curr_poweron'][i_die, :] = np.array(curr_poweron_str.split(';')[1:], dtype='float64')

    return d

    #order of current values in xml file
    # VDDA (A);MinVDDA (A);Max VDDA (A);
    # VDDD (A);MinVDDD (A);Max VDDD (A);
    # VDDDIGPeriph (A);MinVDDDIGPeriph (A);Max VDDDIGPeriph (A);
    # VDD_ANAP (A);MinVDD_ANAP (A);Max VDD_ANAP (A);
    # VDD_ANAN (A);MinVDD_ANAN (A);Max VDD_ANAN (A);
    # Rrefin (A);Vrefin (A);InGaAs_Nref (A);Vcm (A)

def save_stack_raw(d):

    d['im_inj_tile'] = np.zeros((d['N_die'],100 * d['N_gain'], 1032))
    d['im_dark_tile'] = np.zeros((d['N_die'],100 * d['N_gain'], 1032))
    d['im_shortint_tile'] = np.zeros((d['N_die'], 100 * d['N_gain'], 1032))
    d['im_longint_tile'] = np.zeros((d['N_die'], 100 * d['N_gain'], 1032))
    d['im_DC_tile'] = np.zeros((d['N_die'], 100 * d['N_gain'], 1032))

    for i_die in range(d['N_die']):
        for i_gain in range(d['N_gain']):
            d['im_inj_tile'][i_die,i_gain * 100:(i_gain+1) * 100, :] = np.tile((d['im_inj'][i_die,i_gain, :] + 1.5) / 3 * (2 ** 16 - 1), [100, 1])
            d['im_dark_tile'][i_die,i_gain * 100:(i_gain+1) * 100, :] = np.tile((d['im_dark'][i_die,i_gain, :] + 1.5) / 3 * (2 ** 16 - 1), [100, 1])
            d['im_shortint_tile'][i_die, i_gain * 100:(i_gain+1) * 100, :] = np.tile((d['im_shortint'][i_die, i_gain, :] + 1.5) / 3 * (2 ** 16 - 1), [100, 1])
            d['im_longint_tile'][i_die, i_gain * 100:(i_gain+1) * 100, :] = np.tile((d['im_longint'][i_die, i_gain, :] + 1.5) / 3 * (2 ** 16 - 1), [100, 1])

        die = get_RC_from_ind(d, i_die)
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die  + '_sat' +  '.png'),d['im_inj_tile'][i_die,:,:].astype('uint16'))
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die + '_dark' + '.png'), d['im_dark_tile'][i_die,:,:].astype('uint16'))
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die + '_shortint' +  '.png'),d['im_shortint_tile'][i_die,:,:].astype('uint16'))
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die + '_longint' + '.png'), d['im_longint_tile'][i_die,:,:].astype('uint16'))
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die + '_DC' + '.png'),d['im_DC_tile'][i_die, :, :].astype('uint16'))

def save_stack_BP(d):

    d['BP_inj_tile'] = np.zeros((d['N_die'],100 * d['N_gain'], 1032))
    d['BP_dark_tile'] = np.zeros((d['N_die'],100 * d['N_gain'], 1032))
    d['BP_NU_tile'] = np.zeros((d['N_die'], 100 * d['N_gain'], 1032))

    for i_die in range(d['N_die']):
        for i_gain in range(d['N_gain']):
            d['BP_inj_tile'][i_die,i_gain * 100:(i_gain+1) * 100, :] = np.tile(d['BP_inj_gain'][i_die,i_gain, :] *(2**8-1), [100, 1])
            d['BP_dark_tile'][i_die,i_gain * 100:(i_gain+1) * 100, :] = np.tile(d['BP_dark_gain'][i_die,i_gain, :] *(2**8-1), [100, 1])
            d['BP_NU_tile'][i_die, i_gain * 100:(i_gain + 1) * 100, :] = np.tile(d['BP_NU_gain'][i_die, i_gain, :] * (2 ** 8 - 1), [100, 1])

        die = get_RC_from_ind(d, i_die)
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die  + '_BP_inj' +  '.png'),d['BP_inj_tile'][i_die,:,:].astype('uint8'))
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die + '_BP_dark' + '.png'), d['BP_dark_tile'][i_die,:,:].astype('uint8'))
        imwrite(os.path.join(d['dirs_die_sorted'][i_die], die + '_BP_tot' + '.png'),d['BP_NU_tile'][i_die, :, :].astype('uint8'))

def save_plot_dyn_range(d):

    plt.ioff()
    plt.figure(figsize=(15, 10))
    for i_D in range(d['N_die']):
            for i_G in range(d['N_gain']):
                plt.plot(d["im_inj"][i_D,i_G,:])
            for i_G in range(d['N_gain']):
                plt.plot(d["im_dark"][i_D, i_G, :])
        plt.xlim([-10,1042])
        plt.ylim([-2,2])
        plt.title('saturation/dark voltage vs px')
        plt.xlabel('px')
        plt.ylabel('signal, V ')
        legend_inj = ['inj '+g for g in d['list_gain']]
        legend_dark = ['dark '+g for g in d['list_gain']]
        plt.legend(legend_inj+legend_dark)


        die = get_RC_from_ind(d, i_D)

        plt.savefig(os.path.join(d['dirs_die_sorted'][i_D], die + 'v_vs-px.png'),dpi=300)
        plt.show()
        plt.cla()

def throw_on_wafer(d, x):
    # x = d['inj_med'][:,-1]
    wafer = np.full(d['N_r']*d['N_c'],fill_value=-1e-5) #to create contrast on wafer map
    mask = np.ravel(d['map_dirs_die']!=-1) #to know valid dies
    wafer[mask] = x
    wafer = wafer.reshape((d['N_r'],d['N_c']))
    # plt.imshow(wafer)

    return wafer

def plot_on_wafer(d,x,d_key):
    fig, ax = plt.subplots()
    ax_im = ax.imshow(throw_on_wafer(d, x),cmap='RdYlGn')
    for i in range(d['N_r']):  # y
        for ii in range(d['N_c']):  # x
            ax.text(ii, i, d['map_RC_notation'][i, ii], color='k', ha="center", va="center", fontsize='x-small',weight="bold")
    # fig.colorbar(cm.ScalarMappable(), ax=ax)
    # ax_im.set_clim(0, 900)
    # var_name = [k for k,v in d.items() if k == d_key]
    # ax.set_title(var_name[0])
    ax.set_title(d_key)
    ax.grid(which='both')
    ax.set_aspect(aspect=0.25)

    ax.set_xlim((-0.5,3.5))
    ax.xaxis.set_major_locator(IndexLocator(base=2,offset=-1))
    ax.xaxis.set_minor_locator(IndexLocator(base=1,offset=-1))


    ax.set_ylim((23.5,-0.5))
    ax.yaxis.set_major_locator(IndexLocator(base=2,offset=-1))
    ax.yaxis.set_minor_locator(IndexLocator(base=1,offset=-1))

def save_map(d):
    fig, ax = plt.subplots()
    fig.set_size_inches(12, 12)
    ax_im = ax.imshow(throw_on_wafer(d, d['pass_tot']), cmap='RdYlGn')
    for i in range(d['N_r']):  # y
        for ii in range(d['N_c']):  # x
            if d['map_RC_notation'][i, ii] != '':
                i_D  = get_ind_from_RC(d,d['map_RC_notation'][i, ii])
                if d['pass_dyn_range'][i_D] == 1 & d['pass_curr_tot'][i_D] == 1:
                    ax.text(ii, i, d['map_RC_notation'][i, ii] + ', ' + str(int(d['BP_count_NU_tot'][i_D])), color='k', ha="center", va="center", fontsize='small',
                            weight="bold")
                else:
                    ax.text(ii, i, d['map_RC_notation'][i, ii] + ', 1025', color='k', ha="center", va="center", fontsize='small',
                            weight="bold")
    # fig.colorbar(cm.ScalarMappable(), ax=ax)
    # ax_im.set_clim(0, 900)
    # var_name = [k for k,v in d.items() if k == d_key]
    # ax.set_title(var_name[0])
    ax.set_title('pass_tot(green)+#BP'+ '   ' + os.path.basename(d['root']) + '  Y=' + '{:2.2f}'.format(d['yield_tot']) + \
                 '  N_pass = ' + str(np.sum(d["pass_tot"])) + '/'  + str(np.sum(d["N_die"])))
    ax.grid(which='both')
    ax.set_aspect(aspect=0.2)

    ax.set_xticklabels([])
    # ax.set_xlim((-0.5, 13.5))
    ax.xaxis.set_major_locator(IndexLocator(base=2, offset=-1))
    ax.xaxis.set_minor_locator(IndexLocator(base=1, offset=-1))

    ax.set_yticklabels([])
    # ax.set_ylim((16.5, -0.5))
    ax.yaxis.set_major_locator(IndexLocator(base=2, offset=-1))
    ax.yaxis.set_minor_locator(IndexLocator(base=1, offset=-1))

    plt.savefig(os.path.join(d['root'], 'map_' + os.path.basename(d['root']) + '.png'),dpi=300)
    # plt.close()

def get_stat_BP(d):

    d['dark_med'] = np.median(d['im_dark'], axis=2) #average across pixels
    d['inj_med'] = np.median(d['im_inj'], axis=2)
    d['dyn_range'] = d['inj_med'] - d['dark_med']
    d["shortint_med"] = np.median(d['im_shortint'], axis=2)
    d["longint_med"] = np.median(d['im_longint'], axis=2)
    d["DC_med"] = np.median(d['im_DC'], axis=2)

    #process dark
    d["BP_dark_gain"] = np.zeros((d["N_die"],d["N_gain"], 1032), dtype=bool)
    d["BP_dark_tot"] = np.zeros((d["N_die"], 1032), dtype=bool)
    for i in range(d["N_die"]):
        for ii in range(d["N_gain"]):
            low = d["im_dark"][i, ii, 4:1029] > d["dark_med"][i, ii] + d['toler_dark'] * d['dyn_range'][i, ii] #meant in abs values as for positive inj
            high = d["im_dark"][i, ii, 4:1029] < d["dark_med"][i, ii] - d['toler_dark'] * d['dyn_range'][i, ii]
            high_abs = d["im_dark"][i, ii, 4:1029] > -1
            low_abs = d["im_dark"][i, ii, 4:1029] < -1.99
            d["BP_dark_gain"][i,ii,4:1029] = np.array(low | high | high_abs | low_abs)
            d["BP_dark_tot"][i,:] = d["BP_dark_tot"][i,:] | d["BP_dark_gain"][i,ii,:]
    d["BP_count_dark_gain"] = np.sum(d["BP_dark_gain"], axis=2)
    d["BP_count_dark_tot"] = np.sum(d["BP_dark_tot"], axis=1)

    d["BP_inj_gain"] = np.zeros((d["N_die"],d["N_gain"], 1032), dtype=bool)
    d["BP_inj_tot"] = np.zeros((d["N_die"],1032), dtype=bool)
    d["BP_count_inj_low"] = np.zeros((d["N_die"],d["N_gain"]))
    d["BP_count_inj_high"] = np.zeros((d["N_die"],d["N_gain"]))
    for i in range(d["N_die"]):
        for ii in range(d["N_gain"]):
            low = d["im_inj"][i,ii,4:1029] < ( d["inj_med"][i,ii] - d['toler_inj']*d['dyn_range'][i,ii] )
            high = d["im_inj"][i,ii,4:1029] > ( d["inj_med"][i,ii] + d['toler_inj']*d['dyn_range'][i,ii] )
            low_abs = d["im_inj"][i,ii,4:1029] < 1
            high_abs = d["im_inj"][i, ii, 4:1029] > 1.99
            d["BP_inj_gain"][i,ii,4:1029] = np.array(low | high | low_abs | high_abs)
            d["BP_inj_tot"][i, :] = d["BP_inj_gain"][i, ii, :] | d["BP_inj_tot"][i, :]
    d["BP_count_inj_gain"] = np.sum(d["BP_inj_gain"], axis=2)
    d["BP_count_inj_tot"] = np.sum(d["BP_inj_tot"], axis=1)

    d["BP_DC_gain"] = np.zeros((d["N_die"], d["N_gain"], 1032), dtype=bool)
    d["BP_DC_tot"] = np.zeros((d["N_die"], 1032), dtype=bool)
    d['std_DC'] = np.std(d['im_DC'],axis=2)
    for i in range(d['N_die']):
        for ii in range(d["N_gain"]):
            low =  d["im_DC"][i,ii,4:1029] < ( d["DC_med"][i,ii] - 5*d['std_DC'][i,ii] )
            high = d["im_DC"][i,ii,4:1029] > ( d["DC_med"][i,ii] + 5*d['std_DC'][i,ii] )
            d['BP_DC_gain'][i,ii,4:1029] = np.array(low | high)
            d["BP_DC_tot"][i, :] = d["BP_DC_gain"][i, ii, :] | d["BP_DC_tot"][i, :]
    d['BP_count_DC_gain'] = np.sum(d['BP_DC_gain'],axis=2)
    d['BP_count_DC_tot'] = np.sum(d['BP_DC_tot'], axis=1)

    # union of inj and dark over gains
    d['BP_NU_gain'] = np.zeros((d["N_die"], d["N_gain"], 1032))
    for i in range(d["N_die"]):
        for ii in range(d["N_gain"]):
            d['BP_NU_gain'][i,ii,:] = d["BP_dark_gain"][i,ii,:] | d["BP_inj_gain"][i,ii,:] #| d["BP_DC_gain"][i,ii,:]
    d["BP_count_NU_gain"] = np.sum(d["BP_NU_gain"], axis=2)

    #union of inj and dark total
    d['BP_NU_tot'] = np.zeros((d["N_die"], 1032))
    for i in range(d["N_die"]):
        d['BP_NU_tot'][i, :] = d["BP_dark_tot"][i,:] | d["BP_inj_tot"][i,:] #| d["BP_DC_tot"][i,:]
    d["BP_count_NU_tot"] = np.sum(d["BP_NU_tot"], axis=1)

    #fill voids in neighbouring BPs
    for i_D in range(d["N_die"]):
        for i_px in range(4,1030-2,1):
            if d['BP_NU_tot'][i_D,i_px] == 1 and d['BP_NU_tot'][i_D,i_px+1] == 0 and d['BP_NU_tot'][i_D,i_px+2] == 1:
                d['BP_NU_tot'][i_D,i_px+1] = 1

    d["BP_NU_tot_coor"] = []
    for i in range(d["N_die"]):
        d["BP_NU_tot_coor"].append( (np.argwhere(d["BP_NU_tot"][i,:] == 1).squeeze(axis=1)).tolist() )


    d['bin_BP'] = np.array([0, 1, 2, 5, 10, 20, 100, 512, 1024, 1032])
    d['N_bin_BP'] = len(d['bin_BP'])
    d['histo_BP_inj_gain'] = np.zeros((d["N_gain"], d['N_bin_BP'] - 1))
    d['histo_BP_dark_gain'] = np.zeros((d["N_gain"], d['N_bin_BP'] - 1))
    d['histo_BP_DC_gain'] = np.zeros((d["N_gain"], d['N_bin_BP'] - 1))
    for i in range(d["N_gain"]):
        d['histo_BP_dark_gain'][i, :] = np.histogram(d["BP_count_dark_gain"][:,i], bins=d['bin_BP'])[0]
        d['histo_BP_inj_gain'][i, :] = np.histogram(d["BP_count_inj_gain"][:,i], bins=d['bin_BP'])[0]
        d['histo_BP_DC_gain'][i, :] = np.histogram(d["BP_count_DC_gain"][:,i], bins=d['bin_BP'])[0]
    d['histo_BP_dark_tot'] = np.histogram(d["BP_count_dark_tot"], bins=d['bin_BP'])[0]
    d['histo_BP_inj_tot'] = np.histogram(d["BP_count_inj_tot"], bins=d['bin_BP'])[0]
    d['histo_BP_DC_tot'] = np.histogram(d["BP_count_DC_tot"], bins=d['bin_BP'])[0]
    d['histo_BP_NU_tot'] = np.histogram(d["BP_count_NU_tot"], bins=d['bin_BP'])[0]

    return d

def get_stat_cl(d):

    d['bin_cl'] = np.array([0, 1, 2, 5, 10, 20, 100, 512, 1024, 1032])
    d['N_bin_cl'] = len(d['bin_cl'])

    d['cl_map_inj'] = np.zeros((d["N_die"], d["N_gain"], 1032), dtype='uint16')
    d['N_cl_inj'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_1_inj'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_2_inj'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_ge2_inj'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['histo_cl_inj'] = np.zeros((d["N_die"], d["N_gain"], d["N_bin_cl"] - 1))
    cl_size_inj = []
    for i in range(d["N_die"]):
        cl_size_inj_i = []
        for ii in range(d["N_gain"]):
            d['cl_map_inj'][i, ii, :], d['N_cl_inj'][i, ii] = ms.label(d['BP_inj_gain'][i, ii, :])
            cl_size_inj_i.append(ms.sum(d["BP_inj_gain"][i, ii, :], d['cl_map_inj'][i, ii, :],index=range(1, d['N_cl_inj'][i, ii] + 1)))
            d['histo_cl_inj'][i, ii, :], _ = np.histogram(cl_size_inj_i[ii], bins=d['bin_cl'])
            d['N_cl_1_inj'][i, ii] = np.sum(cl_size_inj_i[ii] == 1)
            d['N_cl_2_inj'][i, ii] = np.sum(cl_size_inj_i[ii] == 2)
            d['N_cl_ge2_inj'][i, ii] = np.sum(cl_size_inj_i[ii] >= 2)
        cl_size_inj.append(cl_size_inj_i)
    d['cl_size_inj'] = np.array(cl_size_inj)

    d['cl_map_inj_tot'] = np.zeros((d["N_die"], 1032), dtype='uint16')
    d['N_cl_inj_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_1_inj_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_2_inj_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_ge2_inj_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['histo_cl_inj_tot'] = np.zeros((d["N_die"], d["N_bin_cl"] - 1))
    cl_size_inj_tot = []
    for i in range(d["N_die"]):
        d['cl_map_inj_tot'][i, :], d['N_cl_inj_tot'][i] = ms.label(d['BP_inj_tot'][i,:])
        cl_size_inj_tot.append(ms.sum(d['BP_inj_tot'][i, :], d['cl_map_inj_tot'][i, :],index=range(1, d['N_cl_inj_tot'][i] + 1)))
        d['histo_cl_inj_tot'][i, :], _ = np.histogram(cl_size_inj_tot[i], bins=d['bin_cl'])
        d['N_cl_1_inj_tot'][i] = np.sum(cl_size_inj_tot[i] == 1)
        d['N_cl_2_inj_tot'][i] = np.sum(cl_size_inj_tot[i] == 2)
        d['N_cl_ge2_inj_tot'][i] = np.sum(cl_size_inj_tot[i] >= 2)
    d['cl_size_inj_tot'] = np.array(cl_size_inj_tot)
    d['histo_cl_inj_tot_av'] = np.mean(d['histo_cl_inj_tot'], axis=0)

    d['cl_map_dark'] = np.zeros((d["N_die"], d["N_gain"], 1032), dtype='uint16')
    d['N_cl_dark'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_1_dark'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_2_dark'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_ge2_dark'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['histo_cl_dark'] = np.zeros((d["N_die"], d["N_gain"], d["N_bin_cl"] - 1))
    cl_size_dark = []
    for i in range(d["N_die"]):
        cl_size_dark_i = []
        for ii in range(d["N_gain"]):
            d['cl_map_dark'][i, ii, :], d['N_cl_dark'][i, ii] = ms.label(d['BP_dark_gain'][i, ii, :])
            cl_size_dark_i.append(ms.sum(d["BP_dark_gain"][i, ii, :], d['cl_map_dark'][i, ii, :], index=range(1, d['N_cl_dark'][i, ii] + 1)))
            d['histo_cl_dark'][i, ii, :], _ = np.histogram(cl_size_dark_i[ii], bins=d['bin_cl'])
            d['N_cl_1_dark'][i, ii] = np.sum(cl_size_dark_i[ii] == 1)
            d['N_cl_2_dark'][i, ii] = np.sum(cl_size_dark_i[ii] == 2)
            d['N_cl_ge2_dark'][i, ii] = np.sum(cl_size_dark_i[ii] >= 2)
        cl_size_dark.append(cl_size_dark_i)
    d['cl_size_dark'] = np.array(cl_size_dark)

    d['cl_map_dark_tot'] = np.zeros((d["N_die"], 1032), dtype='uint16')
    d['N_cl_dark_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_1_dark_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_2_dark_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_ge2_dark_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['histo_cl_dark_tot'] = np.zeros((d["N_die"], d["N_bin_cl"] - 1))
    cl_size_dark_tot = []
    for i in range(d["N_die"]):
        d['cl_map_dark_tot'][i, :], d['N_cl_dark_tot'][i] = ms.label(d['BP_dark_tot'][i, :])
        cl_size_dark_tot.append(
            ms.sum(d['BP_dark_tot'][i, :], d['cl_map_dark_tot'][i, :], index=range(1, d['N_cl_dark_tot'][i] + 1)))
        d['histo_cl_dark_tot'][i, :], _ = np.histogram(cl_size_dark_tot[i], bins=d['bin_cl'])
        d['N_cl_1_dark_tot'][i] = np.sum(cl_size_dark_tot[i] == 1)
        d['N_cl_2_dark_tot'][i] = np.sum(cl_size_dark_tot[i] == 2)
        d['N_cl_ge2_dark_tot'][i] = np.sum(cl_size_dark_tot[i] >= 2)
    d['cl_size_dark_tot'] = np.array(cl_size_dark_tot)
    d['histo_cl_dark_tot_av'] = np.mean(d['histo_cl_dark_tot'], axis=0)

    d['cl_map_DC'] = np.zeros((d["N_die"], d["N_gain"], 1032), dtype='uint16')
    d['N_cl_DC'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_1_DC'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_2_DC'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['N_cl_ge2_DC'] = np.zeros((d["N_die"], d["N_gain"]), dtype='uint16')
    d['histo_cl_DC'] = np.zeros((d["N_die"], d["N_gain"], d["N_bin_cl"] - 1))
    cl_size_DC = []
    for i in range(d["N_die"]):
        cl_size_DC_i = []
        for ii in range(d["N_gain"]):
            d['cl_map_DC'][i, ii, :], d['N_cl_DC'][i, ii] = ms.label(d['BP_DC_gain'][i, ii, :])
            cl_size_DC_i.append(ms.sum(d["BP_DC_gain"][i, ii, :], d['cl_map_DC'][i, ii, :], index=range(1, d['N_cl_DC'][i, ii] + 1)))
            d['histo_cl_DC'][i, ii, :], _ = np.histogram(cl_size_DC_i[ii], bins=d['bin_cl'])
            d['N_cl_1_DC'][i, ii] = np.sum(cl_size_DC_i[ii] == 1)
            d['N_cl_2_DC'][i, ii] = np.sum(cl_size_DC_i[ii] == 2)
            d['N_cl_ge2_DC'][i, ii] = np.sum(cl_size_DC_i[ii] >= 2)
        cl_size_DC.append(cl_size_DC_i)
    d['cl_size_DC'] = np.array(cl_size_DC)

    d['cl_map_DC_tot'] = np.zeros((d["N_die"], 1032), dtype='uint16')
    d['N_cl_DC_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_1_DC_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_2_DC_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_ge2_DC_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['histo_cl_DC_tot'] = np.zeros((d["N_die"], d["N_bin_cl"] - 1))
    cl_size_DC_tot = []
    for i in range(d["N_die"]):
        d['cl_map_DC_tot'][i, :], d['N_cl_DC_tot'][i] = ms.label(d['BP_DC_tot'][i, :])
        cl_size_DC_tot.append(
            ms.sum(d['BP_DC_tot'][i, :], d['cl_map_DC_tot'][i, :], index=range(1, d['N_cl_DC_tot'][i] + 1)))
        d['histo_cl_DC_tot'][i, :], _ = np.histogram(cl_size_DC_tot[i], bins=d['bin_cl'])
        d['N_cl_1_DC_tot'][i] = np.sum(cl_size_DC_tot[i] == 1)
        d['N_cl_2_DC_tot'][i] = np.sum(cl_size_DC_tot[i] == 2)
        d['N_cl_ge2_DC_tot'][i] = np.sum(cl_size_DC_tot[i] >= 2)
    d['cl_size_DC_tot'] = np.array(cl_size_DC_tot)
    d['histo_cl_DC_tot_av'] = np.mean(d['histo_cl_DC_tot'], axis=0)

    d['cl_map_NU_tot'] = np.zeros((d["N_die"], 1032), dtype='uint16')
    d['N_cl_NU_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_1_NU_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_2_NU_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['N_cl_ge2_NU_tot'] = np.zeros(d["N_die"], dtype='uint16')
    d['histo_cl_NU_tot'] = np.zeros((d["N_die"], d["N_bin_cl"] - 1))
    cl_size_NU_tot = []
    for i in range(d["N_die"]):
        d['cl_map_NU_tot'][i, :], d['N_cl_NU_tot'][i] = ms.label(d['BP_NU_tot'][i, :])
        cl_size_NU_tot.append(
            ms.sum(d['BP_NU_tot'][i, :], d['cl_map_NU_tot'][i, :], index=range(1, d['N_cl_NU_tot'][i] + 1)))
        d['histo_cl_NU_tot'][i, :], _ = np.histogram(cl_size_NU_tot[i], bins=d['bin_cl'])
        d['N_cl_1_NU_tot'][i] = np.sum(cl_size_NU_tot[i] == 1)
        d['N_cl_2_NU_tot'][i] = np.sum(cl_size_NU_tot[i] == 2)
        d['N_cl_ge2_NU_tot'][i] = np.sum(cl_size_NU_tot[i] >= 2)
    d['cl_size_NU_tot'] = np.array(cl_size_NU_tot)
    d['histo_cl_NU_tot_av'] = np.mean(d['histo_cl_NU_tot'], axis=0)



    # plt.figure()
    # plt.plot(d['N_cl_1_dark_tot'])

    return d

def get_stat_curr(d):

    #range of typical values of currents
    d['bin_curr_VDDA'] = np.arange(0, 1, 1e-2)  # VDDA, 0
    d['bin_curr_VDDD'] = np.arange(0, 5e-3, 2e-5)  # VDDD, 3
    # d['bin_curr'] = np.arange(0, 0.2, 0.5e-2)  # ANAP, 8
    # d['bin_curr'] = np.arange(0, 0.2, 0.5e-2)  # Rrefin, 15
    # d['bin_curr'] = np.arange(0, 0.2, 0.5e-2)  # Vcm, 18

    d['histo_curr_running_VDDA'] = np.histogram(d['curr_running'][:, 0], bins=d['bin_curr_VDDA'])[0] / d['N_die']
    d['histo_curr_running_VDDD'] = np.histogram(d['curr_running'][:, 5], bins=d['bin_curr_VDDD'])[0] / d['N_die']
    # d['histo_curr_poweron'], b = np.histogram(np.log(np.abs(d['curr_poweron'][:, 0])), bins=20)
    # d['histo_curr_CBGB'], b = np.histogram(d['curr_CBGB'][:, 0], bins=d['bin_curr'])

    # plt.plot(b[:-1], d['histo_curr_poweron'])
    # plt.plot(b[:-1], d['histo_curr_CBGB'])

    return d

def spec_check(d):
    d['pass_dyn_range'] = np.zeros(d['N_die'], dtype=bool)
    for i in range(d['N_die']):
        d['pass_dyn_range'][i] = np.all(np.array(d['dyn_range'][i,:] > 2)) #gain_av > 2
    d['yield_dyn_range'] = np.sum(d['pass_dyn_range']) / d['N_die']

    d['pass_inj_min'] = np.zeros(d['N_die'],dtype=bool)
    for i in range(d['N_die']):
        d['pass_inj_min'][i] = np.all(np.array(d['inj_med'][i,:] > 1)) #for all gains
    d['yield_inj_min'] = np.sum(d['pass_inj_min']) / d['N_die']

    d['pass_dark_min'] = np.zeros(d['N_die'],dtype=bool)
    for i in range(d['N_die']):
        d['pass_dark_min'][i] = np.all(np.array(d['dark_med'][i,:] < -1))
    d['yield_dark_min'] = np.sum(d['pass_dark_min']) / d['N_die']

    d['pass_BP_oper'] = np.array(d['BP_count_NU_tot'] <= 6)
    d['yield_BP_oper'] = np.sum(d['pass_BP_oper']) / d['N_die']

    # d['pass_cl_inj_1'] = np.array(d['N_cl_inj_1_tot'] <= 6)
    # d['yield_cl_inj_1'] = np.sum(d['pass_cl_inj_1']) / d['N_die']
    # d['pass_cl_inj_2'] = np.array(d['N_cl_inj_2_tot'] == 0)
    # d['yield_cl_inj_2'] = np.sum(d['pass_cl_inj_2']) / d['N_die']
    # d['pass_cl_inj_ge2'] = np.array(d['N_cl_inj_ge2_tot'] == 0)
    # d['yield_cl_inj_ge2'] = np.sum(d['pass_cl_inj_ge2']) / d['N_die']
    # d['pass_cl_inj'] = d['pass_cl_inj_1'] & d['pass_cl_inj_2'] & d['pass_cl_inj_ge2']
    # d['yield_cl_inj'] = np.sum(d['pass_cl_inj']) / d['N_die']

    d['pass_cl_ge2_NU_tot'] = np.array(d['N_cl_ge2_NU_tot'] == 0)
    d['yield_cl_ge2_NU_tot'] = np.sum(d['pass_cl_ge2_NU_tot']) / d['N_die']

    d['pass_curr_VDDA'] = np.array(d['curr_running'][:,0] > 0) & np.array(d['curr_running'][:,0] <= 0.13) #min, max
    d['yield_curr_VDDA'] = np.sum(d['pass_curr_VDDA']) / d['N_die']
    d['pass_curr_VDDD'] = np.array(d['curr_running'][:,3] < 0.07)
    d['yield_curr_VDDD'] = np.sum(d['pass_curr_VDDD']) / d['N_die']
    d['pass_curr_VCM'] = np.array(d['curr_running'][:, 18] > 0.145) & np.array(d['curr_running'][:, 18] < 0.155)
    d['yield_curr_VCM'] = np.sum(d['pass_curr_VCM']) / d['N_die']

    d['pass_dyn_range_tot'] = d['pass_inj_min'] & d['pass_dark_min'] & d['pass_dyn_range']
    d['pass_curr_tot'] = d['pass_curr_VDDA'] & d['pass_curr_VDDD'] & d['pass_curr_VCM']
    d['pass_tot'] = d['pass_dyn_range_tot'] & d['pass_BP_oper'] & d['pass_cl_ge2_NU_tot'] & d['pass_curr_tot']
    d['yield_tot'] = np.sum(d['pass_tot']) / d['N_die']

    return d

def get_RC_from_ind(d, ind):
    coor = np.where(d['map_dirs_die'] == ind)
    return d['map_RC_notation'][coor][0]

def get_ind_from_RC(d, RC):
    coor = np.where(d['map_RC_notation'] == RC)
    return d['map_dirs_die'][coor][0]

def log_excel(d):
    sheet = 'Sheet1'
    wb = openpyxl.load_workbook(d['xfile_template'])
    xsheet = wb[sheet]
    str_gain = ['30A','30B','105A','105B','205A','205B','505A','505B']

    for i in range(d['N_die']):
        xsheet['A' + str(10*i + 2)] = str(get_RC_from_ind(d, i))
        xsheet['B' + str(10 * i + 2)] = 'tot'
        xsheet['F' + str(10 * i + 2)] = d["BP_count_NU_tot"][i]
        xsheet['G' + str(10 * i + 2)] = str(d["BP_NU_tot_coor"][i])
        xsheet['H' + str(10*i + 2)] = d['curr_running'][i,0]
        xsheet['I' + str(10*i + 2)] = d['curr_running'][i, 3]
        xsheet['J' + str(10*i + 2)] = d['pass_inj_min'][i]
        xsheet['K' + str(10*i + 2)] = d['pass_dark_min'][i]
        xsheet['L' + str(10*i + 2)] = d['pass_dyn_range'][i]
        xsheet['M' + str(10*i + 2)] = d['pass_BP_oper'][i]
        xsheet['N' + str(10 * i + 2)] = d['pass_cl_ge2_NU_tot'][i]
        xsheet['O' + str(10*i + 2)] = d['pass_curr_VDDA'][i]
        xsheet['P' + str(10*i + 2)] = d['pass_curr_VDDD'][i]
        xsheet['Q' + str(10*i + 2)] = d['pass_tot'][i]


        for ii in range(d['N_gain']):
            xsheet['B' + str(10*i + ii + 1 + 2)] = str_gain[ii]
            xsheet['C' + str(10*i + ii + 1 + 2)] = d["inj_med"][i, ii]
            xsheet['D' + str(10*i + ii + 1 + 2)] = d["dark_med"][i,ii]
            xsheet['E' + str(10*i + ii + 1 + 2)] = d["dyn_range"][i, ii]
            xsheet['F' + str(10*i + ii + 1 + 2)] = d["BP_count_NU_gain"][i, ii]

    offset = 78*10 + 2 + 5
    for i in range(d['N_die']):
        coor = np.where(d['map_dirs_die'] == i)
        xsheet.cell(row=coor[0][0] + offset, column=2 + coor[1][0]).value = d['pass_tot'][i]
        # if d['pass_tot'][i] == True:
        #     xsheet.cell(row=coor[0][0]+offset, column=2+coor[1][0]).value = 'pass'
        # else:
        #     xsheet.cell(row=coor[0][0]+offset, column=2+coor[1][0]).value = 'fail'

    xsheet['B813'] = d["yield_inj_min"]
    xsheet['B814'] = d["yield_dark_min"]
    xsheet['B815'] = d["yield_dyn_range"]
    xsheet['B816'] = d["yield_BP_oper"]
    xsheet['B817'] = d["yield_curr_VDDA"]
    xsheet['B818'] = d["yield_curr_VDDD"]
    xsheet['B819'] = d["yield_tot"]
    xsheet['B820'] = np.sum(d["pass_tot"])

    d['xfile'] = os.path.join(d['root'], os.path.basename(d['root']) + '.xlsx')
    wb.save(d['xfile'])

def get_analysis(d):

    d = D['waf'][0]

    plt.figure()
    for i_G in range(d['N_gain']):
        plt.plot(d['dyn_range'][:, i_G],label=str(i_G))
    plt.legend()

    plt.figure()
    plt.plot(d['pass_curr_tot'],label='pass curr tot')
    plt.plot(d['pass_BP_oper'] + 1, label='pass BP_oper')
    plt.plot(d['pass_inj_min'] + 4,label='pass inj min')
    plt.plot(d['pass_dark_min'] + 4, label='pass dark min')
    plt.plot(d['dyn_range'][:, 2] + 4 ,label='dyn range')
    plt.legend()

    plt.figure()
    plt.plot(d['histo_BP_NU_tot'])

    plt.figure()
    plt.plot(d['N_cl_1_NU_tot'],label='cl 1',color='k')
    plt.plot(d['N_cl_2_NU_tot'],label='cl 2',color='r')
    plt.plot(d['N_cl_ge2_NU_tot'],label='cl ge2',color='b')
    plt.legend()
    plt.ylim([-1, 5])

    BP1 = np.zeros(d['N_gain'])
    for i_D in range(d['N_die']):
        for i_G in range(d['N_gain']):
            if d['pass_dyn_range'][i_D] == 1:
                if np.any(d['BP_NU_gain'][i_D,i_G,:5]) == 1:
                   BP1[i_G] += 1

    plt.figure()
    for i_G in range(d['N_gain']):
        plt.plot(np.sum(d['BP_DC_gain'][:,i_G,:],axis=0),label=str(i_G))
    plt.legend()

    plt.figure()
    for i_G in range(d['N_gain']):
        plt.plot(np.sum(d['im_DC'][:,i_G,:],axis=0),label=str(i_G))
    plt.legend()

    plt.figure()
    plt.imshow(d['BP_count_DC_gain'][d['pass_tot'],:].T,vmin=0,vmax=5)


    plt.figure()
    plt.imshow(d['BP_inj_tot'][:,:])

    plt.figure()
    plt.imshow(d["BP_count_inj_gain"].T,vmin=0,vmax=5)

    plt.figure()
    plt.plot(d["BP_count_inj_tot"],label='inj_tot')
    plt.plot(d["BP_count_DC_tot"],label='DC_tot')
    plt.legend()

    plt.figure()
    for i_G in range(d['N_gain']):
        plt.plot(d['histo_BP_inj_gain'][i_G,:],label=str(i_G))
    plt.legend()

    plt.figure()
    plt.plot(d['N_cl_1_DC_tot'])

    i_G = 4
    plt.figure()
    plt.errorbar(x = np.array(range(d['N_die'])), y = d["DC_med"][:,i_G],yerr=d['std_DC'][:,i_G])


    plt.figure()
    for i_D in range(d['N_die']):
        for i_G in range(d['N_gain']):
            if d['pass_tot'][i_D] ==1:
                plt.plot(d["im_inj"][i_D,5,:])
                plt.plot(d["im_dark"][i_D, 5, :])

    plt.figure()
    for i_D in range(d['N_die']):
        for i_G in range(d['N_gain']):
            if d['pass_tot'][i_D] == 1:
                plt.plot(d["BP_inj_gain"][i_D, i_G, :])
                # plt.plot(d["im_dark"][i_D, 1, :])

                # plt.plot(d["im_shortint"][i_D, i_G, :])
                # plt.plot(d["im_longint"][i_D, i_G, :])

                # plt.plot(d["im_DC"][i_D, 4, :])

    # plt.figure()
    # # plt.imshow(d["im_dark"][d['pass_tot'],1,:],vmin=-1.5,vmax=-1)
    # plt.imshow(d["im_dark"][:,5,:],vmin=-1.5,vmax=-1)
    # # plt.imshow(d["im_inj"][:,5,:],vmin=1,vmax=1.5)

    plt.figure()
    for i_D in range(d['N_die']):
        if d['pass_tot'][i_D] == 1:
            plt.plot(d["im_DC"][i_D,1,:])

    inj_low = np.amin(np.array(d['im_inj'][:, :, 4:1029]), axis=2)
    inj_high = np.amax(np.array(d['im_inj'][:, :, 4:1029]), axis=2)
    dark_high = np.amax(np.array(d['im_dark'][:, :, 4:1029]),axis=2)
    dark_low = np.amin(np.array(d['im_dark'][:, :, 4:1029]), axis=2)
    plt.figure()
    # plt.plot(d['pass_dark_min'])
    for i_G in range(d['N_gain']):
        # plt.plot(d['dyn_range'][:,i_G])
        plt.plot(d['inj_med'])
        plt.plot(d['dark_med'])

    plt.figure()
    plt.plot(d["BP_inj_tot_av"])

    d["BP_inj_gain_av"] = np.sum(d["BP_inj_gain"][d['pass_dyn_range_tot']], axis=0) #over dies, px and gains remain; only for passed dies
    d["BP_inj_tot_av"] = np.sum(d["BP_inj_tot"][d['pass_dyn_range_tot']], axis=0)

    d["BP_dark_gain_av"] = np.sum(d["BP_dark_gain"][d['pass_dyn_range_tot']], axis=0)
    d["BP_dark_tot_av"] = np.sum(d["BP_dark_tot"][d['pass_dyn_range_tot']], axis=0)
    plt.plot(d["BP_dark_tot_av"])

def proc_batch(D):
    D['waf'] = []
    D['N_waf'] = len(D['root_list'])

    for i in range(len(D['root_list'])):
        i=0
        d = {}
        d['list_gain'] = [
            "*CFB_30_bankA*",
            "*CFB_30_bankB*",
            "*CFB_105_bankA*",
            "*CFB_105_bankB*",
            "*CFB_205_bankA*",
            "*CFB_205_bankB*",
            "*CFB_505_bankA*",
            "*CFB_505_bankB*",
        ]  # all gains of XRO3508
        d['N_gain'] = len(d['list_gain'])
        d['N_bank'] = int(d['N_gain'] / 2)  # bank B used for IWR mode
        d['toler_dark'] = 0.1  # tolerance factors for NU
        d['toler_inj'] = 0.2

        d['root'] = D['root_list'][i]
        d['dirs_die'] = [f for f in glob.glob(os.path.join(d['root'], 'R??-C??'))]
        d['xfile_template'] = r'c:\Users\michael.hakl\Documents\XRO3508\3508_template.xlsx'

        d = get_map_wafer(d)
        d = load_im(d)
        d = load_curr(d)

        d = get_stat_BP(d)
        d = get_stat_cl(d)
        d = get_stat_curr(d)

        d = spec_check(d)

        save_stack_raw(d)
        save_stack_BP(d)
        log_excel(d)
        save_map(d)
        D['waf'].append(d)

    return D

def stat_batch(D):
    D['waf_yield_tot'] = np.zeros(D['N_waf'])
    D['waf_N_pass_tot'] = np.zeros(D['N_waf'])
    for i in range(D['N_waf']):
        D['waf_yield_tot'][i] = D['waf'][i]['yield_tot']
        D['waf_N_pass_tot'][i] = np.sum(D['waf'][i]['pass_tot'])

    return D

D = proc_batch(D)
D = stat_batch(D)

d = D['waf'][0]

# plot_on_wafer(D['waf'][0],D['waf'][0]['pass_tot'],'pass_tot')
# plot_on_wafer(D['waf'][0],D['waf'][0]['pass_dyn_range'], 'pass_dyn_range')
# plot_on_wafer(D['waf'][0],D['waf'][0]['pass_BP_oper'], 'pass_BP_oper')
# # plot_on_wafer(D['waf'][0],D['waf'][0]['pass_curr_tot'], 'pass_curr_tot')
# # plot_on_wafer(D['waf'][0],D['waf'][0]['pass_cl_inj'], 'pass_cl_inj')
# plot_on_wafer(D['waf'][0],D['waf'][0]['BP_count_NU_tot'], 'BP_count_NU_tot')


# get_ind_from_RC(d, 'R13C01')
# get_RC_from_ind(d, 76)


#### one wafer proc

# d['root'] = D['root_list'][0]
# xfile = D['xfile_list'][0]

# d['dirs_die'] = np.array([f for f in glob.glob(os.path.join(d['root'],'R??-C??'))])
# d['list_gain'] = [
#     "*CFB_30_bankA*",
#     "*CFB_30_bankB*",
#     "*CFB_105_bankA*",
#     "*CFB_105_bankB*",
#     "*CFB_205_bankA*",
#     "*CFB_205_bankB*",
#     "*CFB_505_bankA*",
#     "*CFB_505_bankB*",
# ] #all gains of XRO3508
# d['N_gain'] = len(d['list_gain'])
# d['N_bank'] = int(d['N_gain']/2) # bank B used for IWR mode
# d['toler_dark'] = 0.1 #tolerance factors for NU
# d['toler_inj'] = 0.2

# d = get_map_wafer(d)
# d = load_im(d)
# d = load_curr(d)
#
# d = get_stat_BP(d)
# d = get_stat_cl(d)
# d = get_stat_curr(d)
#
# d = spec_check(d)
#
# save_stack_raw(d)
# save_stack_BP(d)
# log_excel(d, xfile)
#
# plot_on_wafer(d, d['pass_tot'],'pass_tot')
# plot_on_wafer(d,d['pass_dyn_range'], 'pass_dyn_range')
# plot_on_wafer(d,d['pass_BP_oper'], 'pass_BP_oper')
# # plot_on_wafer(d,d['pass_curr_tot'], 'pass_curr_tot')
# # plot_on_wafer(d,d['pass_cl_inj'], 'pass_cl_inj')
# plot_on_wafer(d,d['BP_count_NU_tot'], 'BP_count_NU_tot')
#
# # plot_on_wafer(d,d['pass_inj_min'], 'pass_inj_min')
# # plot_on_wafer(d,d['pass_dark_min'], 'pass_dark_min')
# # plot_on_wafer(d,d['pass_BP_DC'], 'pass_BP_DC')
# #
# # plot_on_wafer(d,d['curr_poweron'][:,0], 'curr poweron')
# # plot_on_wafer(d,d['curr_CBGB'][:,0], 'curr CBGB')
# # plot_on_wafer(d,d['curr_running'][:,0], 'curr running')
# #
# #
# plt.figure()
# plt.plot(np.mean(d['dyn_range'],axis=1))
# # plt.plot(d['curr_running'][:,0]*20)
# # #
# plt.plot(np.mean(d['dark_med'],axis=1))
# plt.plot(np.mean(d['inj_med'],axis=1))
# #
# plt.figure()
# for i in range(d['N_gain']):
#     plt.plot(d["BP_dark_gain"][53,i,:])
# plt.plot(d["BP_dark_tot"][53,:])
# plt.plot(d['BP_NU_tot'][53])
#
# # plt.plot(d['BP_NU_tot'][27,:])
# plt.plot(d['BP_count_DC_tot'])
# plt.plot(d['BP_count_NU_tot'])
# plt.plot(d['N_cl_ge2_inj_tot'])
# plt.plot(d['N_cl_2_inj_tot'])
# plt.errorbar(x=np.arange(d['N_die']),y=np.mean(d['DC_med'],axis=1),yerr=np.mean(d['std_DC'],axis=1))
# #
# plt.figure()
# plt.plot(d['im_inj'][74,0,:])
# plt.plot(d['im_dark'][74,0,:])
# plt.plot(d['im_DC'][74,0,:])


# plt.figure()
# plt.plot(d['histo_BP_inj_tot'])
# plt.plot(d['pass_dyn_range_tot'])
# plt.plot(d['pass_BP_oper'])
# plt.plot(d['curr_running'][:,18])
# plt.plot(d['pass_dyn_range_tot'] & d['pass_BP_oper'] & d['pass_curr_tot'])
# plt.plot(d['pass_tot'])
# plt.plot(d['pass_curr_VCM'])


#mycomment2



